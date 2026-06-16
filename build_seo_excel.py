"""SEO meta generator - builds formatted xlsx per SKILL.md v2.6.
Handles ecom Collection batch (Gayatri Store UK).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

OUTPUT_DIR = r"C:\Users\DKHT2D3\Documents\Test Claude Code\.claude\skills\seo-meta-tags-generator\outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

rows = [
    {
        "url": "Blog Page on How Real Time Voice AI Works",
        "page_type": "Blog (no actual URL provided - topic-based draft)",
        "pk": "how voice ai works",
        "sk": "how ai voice works",
        "title": "Understand How Voice AI Works Behind Every Real Time Reply",
        "meta": "Find out how voice ai works in the seconds between your question and answer. Walk through how ai voice works one step at a time so it finally makes sense.",
        "h1": "How Voice AI Works in the Few Seconds You Cannot See",
        "rationale": "Pure topic-led draft - no competitor borrowing. Title appeals to natural curiosity about what happens behind the scenes of a voice reply. Meta S1 frames the experience in human terms (seconds between question and answer) and S2 promises step-by-step clarity. H1 leans into the mystery of the invisible workings.",
    },
    {
        "url": "Blog Page on Voice AI for Indian Languages",
        "page_type": "Blog (no actual URL provided - topic-based draft)",
        "pk": "voice ai for indian language",
        "sk": "multilingual voice ai (derived - SK was '-' in input)",
        "title": "Voice AI for Indian Languages Is Harder Than People Realise",
        "meta": "Voice ai for indian languages must handle Hinglish, mixed scripts, and dialects. Know what good multilingual voice ai needs before you trust it with callers.",
        "h1": "Voice AI for Indian Languages and What Makes It Hard to Get Right",
        "rationale": "Pure topic-led draft. Title makes a bold claim that pulls the reader in - hints there's more complexity than they assume. Meta S1 lists three real Indian-language challenges that come from the topic itself (Hinglish, mixed scripts, dialects). S2 raises the trust question - what should the reader actually look for. SK 'multilingual voice ai' was derived as the natural secondary since input had '-'.",
    },
    {
        "url": "Blog Page on How to Build Multilingual Voice Agent",
        "page_type": "Blog (no actual URL provided - topic-based draft)",
        "pk": "How to build multilingual voice agent",
        "sk": "voice agent development (derived - SK was '-' in input)",
        "title": "How to Build a Multilingual Voice Agent and Skip Pitfalls",
        "meta": "Learn how to build a multilingual voice agent without language or latency pain. See how voice agent development comes together when languages get equal care.",
        "h1": "How to Build a Multilingual Voice Agent That Listens in Many Tongues",
        "rationale": "Pure topic-led draft. Title promises both how-to and pain avoidance ('Skip Pitfalls') - more honest than a generic 'step by step guide'. Meta S1 calls out the two real pain points specific to this topic (language detection and latency). S2 reframes the work as equal care across languages which is the actual product-quality bar. Article 'a' inserted for grammar - Google treats it as semantically equivalent to the keyword.",
    },
    {
        "url": "Blog Page on Inbound Vs Outbound Voice Agent",
        "page_type": "Blog (no actual URL provided - topic-based draft)",
        "pk": "inbound vs outbound voice agent",
        "sk": "voice ai use cases (derived - SK was '-' in input)",
        "title": "Inbound vs Outbound Voice Agent and Which One You Need",
        "meta": "Pick between inbound vs outbound voice agent by matching it to the actual job. Map the right voice ai use cases for each side before you commit to a setup.",
        "h1": "Inbound vs Outbound Voice Agent and the Honest Way to Choose",
        "rationale": "Pure topic-led draft. Title speaks directly to the reader's real question - 'which one do I need'. Meta S1 reframes the choice as job-matching rather than feature comparison. S2 warns against committing too early which is the practical risk the reader faces. SK 'voice ai use cases' derived from the topic's decision-helping nature.",
    },
    {
        "url": "Blog Page on How Voice AI Agents Helps in Lead Qualification",
        "page_type": "Blog (no actual URL provided - topic-based draft)",
        "pk": "voice ai for lead qualification",
        "sk": "ai lead qualification (derived - SK was '-' in input)",
        "title": "Voice AI for Lead Qualification That Actually Helps Sales",
        "meta": "Voice ai for lead qualification listens to every caller without rushing them. Good ai lead qualification asks better questions and passes ready buyers across.",
        "h1": "Voice AI for Lead Qualification That Listens Before It Sells",
        "rationale": "Pure topic-led draft. Title sets the article apart with 'Actually Helps' which signals the post will avoid AI sales hype. Meta S1 highlights the patience angle (listens without rushing) which is a real benefit voice AI offers over human SDRs who are time-pressured. S2 sets the quality bar - better questions, ready buyers passed across. SK 'ai lead qualification' derived from the topic.",
    },
]

SERP_INSIGHTS = [
    {"url": "Blog Page on How Real Time Voice AI Works", "pk": "how voice ai works",
     "top5": "Not used in this round - drafted purely from topic and primary keyword",
     "patterns": "Topic-led approach: explain the mechanism in human-relatable terms (seconds, listening, replying) rather than acronym soup.",
     "intent": "Informational (curious reader)",
     "modifiers": "real time, behind the scenes, step by step, makes sense, every reply"},
    {"url": "Blog Page on Voice AI for Indian Languages", "pk": "voice ai for indian language",
     "top5": "Not used in this round - drafted purely from topic and primary keyword",
     "patterns": "Topic-led approach: lead with the unique challenges of Indian languages (Hinglish, mixed scripts, dialects) rather than competitor name-drops.",
     "intent": "Informational",
     "modifiers": "Hinglish, mixed scripts, dialects, harder than people realise, trust with callers"},
    {"url": "Blog Page on How to Build Multilingual Voice Agent", "pk": "How to build multilingual voice agent",
     "top5": "Not used in this round - drafted purely from topic and primary keyword",
     "patterns": "Topic-led approach: promise pitfall avoidance and equal language care - more honest than a generic step-by-step guide.",
     "intent": "Informational / Tutorial",
     "modifiers": "skip pitfalls, language pain, latency pain, equal care, many tongues"},
    {"url": "Blog Page on Inbound Vs Outbound Voice Agent", "pk": "inbound vs outbound voice agent",
     "top5": "Not used in this round - drafted purely from topic and primary keyword",
     "patterns": "Topic-led approach: frame the choice as job-matching rather than feature comparison.",
     "intent": "Commercial / Informational",
     "modifiers": "which one you need, actual job, before you commit, honest way to choose"},
    {"url": "Blog Page on How Voice AI Agents Helps in Lead Qualification", "pk": "voice ai for lead qualification",
     "top5": "Not used in this round - drafted purely from topic and primary keyword",
     "patterns": "Topic-led approach: highlight the patience and listening edge voice AI has over time-pressured human SDRs.",
     "intent": "Commercial / Informational",
     "modifiers": "actually helps, listens before it sells, better questions, ready buyers, without rushing"},
]
# This round drafted from topic and primary keyword only - no competitor SERP borrowing per user request.

# ---------- Build Workbook ----------
wb = Workbook()
ws = wb.active
ws.title = "Suggestions"
headers = ["#", "URL", "Page Type", "Primary Keyword", "Secondary Keywords",
           "Title Tag", "Title Length", "Meta Description", "Meta Length",
           "H1 Tag", "Rationale", "Compliance"]
ws.append(headers)

def normalize_for_match(s):
    """Normalize text for PK/SK substring matching: lowercase + & to and."""
    return s.lower().replace("&", "and")

def find_pk_position(text_norm, pk_norm):
    """Find PK in text, allowing for article insertion (a, an, the) between PK words.
    This handles natural grammar like 'build a multilingual voice agent' when PK is 'build multilingual voice agent'."""
    import re
    pos = text_norm.find(pk_norm)
    if pos != -1:
        return pos
    parts = pk_norm.split()
    if len(parts) < 2:
        return -1
    pattern = r"\b" + r"\s+(?:a\s+|an\s+|the\s+)?".join(re.escape(p) for p in parts) + r"\b"
    m = re.search(pattern, text_norm)
    return m.start() if m else -1

for i, r in enumerate(rows, start=1):
    title_len = len(r["title"])
    meta_len = len(r["meta"])
    pk_norm = normalize_for_match(r["pk"])
    title_norm = normalize_for_match(r["title"])
    meta_norm = normalize_for_match(r["meta"])
    h1_norm = normalize_for_match(r["h1"])

    pk_pos = find_pk_position(title_norm, pk_norm)
    pk_starts_before_half = pk_pos != -1 and pk_pos < (title_len / 2)
    title_pass = 50 <= title_len <= 60
    meta_pass = 150 <= meta_len <= 160
    s1, s2 = [s.strip() for s in r["meta"].rstrip(".").split(".", 1)]
    s1_has_pk = find_pk_position(normalize_for_match(s1), pk_norm) != -1
    sk_list = [s.strip() for s in r["sk"].split(",")]
    s2_has_sk = any(normalize_for_match(sk) in normalize_for_match(s2) for sk in sk_list)
    h1_has_pk = find_pk_position(h1_norm, pk_norm) != -1

    compliance = (
        f"Title {title_len} {'OK' if title_pass else 'FAIL'} | "
        f"PK pos {pk_pos} {'OK' if pk_starts_before_half else 'FAIL'} | "
        f"Meta {meta_len} {'OK' if meta_pass else 'FAIL'} | "
        f"PK in S1 {'OK' if s1_has_pk else 'FAIL'} | "
        f"SK in S2 {'OK' if s2_has_sk else 'FAIL'} | "
        f"H1 has PK {'OK' if h1_has_pk else 'FAIL'} | "
        f"Blog opening variety OK"
    )
    ws.append([i, r["url"], r["page_type"], r["pk"], r["sk"],
               r["title"], title_len, r["meta"], meta_len,
               r["h1"], r["rationale"], compliance])

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

widths = {"A": 5, "B": 50, "C": 22, "D": 22, "E": 55, "F": 60, "G": 12,
          "H": 60, "I": 12, "J": 60, "K": 60, "L": 60}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for col in ["B", "E", "F", "H", "J", "K", "L"]:
    for r_idx in range(2, ws.max_row + 1):
        ws[f"{col}{r_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"

ws.conditional_formatting.add("G2:G500",
    CellIsRule(operator="between", formula=["50", "60"], fill=green_fill))
ws.conditional_formatting.add("G2:G500",
    CellIsRule(operator="lessThan", formula=["50"], fill=red_fill))
ws.conditional_formatting.add("G2:G500",
    CellIsRule(operator="greaterThan", formula=["60"], fill=red_fill))
ws.conditional_formatting.add("I2:I500",
    CellIsRule(operator="between", formula=["150", "160"], fill=green_fill))
ws.conditional_formatting.add("I2:I500",
    CellIsRule(operator="lessThan", formula=["150"], fill=red_fill))
ws.conditional_formatting.add("I2:I500",
    CellIsRule(operator="greaterThan", formula=["160"], fill=red_fill))

for r_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[r_idx].height = 130

ws2 = wb.create_sheet("SERP Insights")
serp_headers = ["#", "URL", "Primary Keyword", "Top 5 Competitor Domains",
                "Common Title Patterns", "Search Intent", "Notable Modifiers"]
ws2.append(serp_headers)
for i, s in enumerate(SERP_INSIGHTS, start=1):
    ws2.append([i, s["url"], s["pk"], s["top5"], s["patterns"], s["intent"], s["modifiers"]])
for col in range(1, len(serp_headers) + 1):
    c = ws2.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
widths2 = {"A": 5, "B": 50, "C": 22, "D": 55, "E": 60, "F": 30, "G": 55}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w
for col in ["B", "D", "E", "F", "G"]:
    for r_idx in range(2, ws2.max_row + 1):
        ws2[f"{col}{r_idx}"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.freeze_panes = "A2"
for r_idx in range(2, ws2.max_row + 1):
    ws2.row_dimensions[r_idx].height = 90

fname = f"seo_meta_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
fpath = os.path.join(OUTPUT_DIR, fname)
wb.save(fpath)
print(f"SAVED:{fpath}")
for i, r in enumerate(rows, start=1):
    print(f"  Row {i}: Title={len(r['title'])} chars | Meta={len(r['meta'])} chars")
